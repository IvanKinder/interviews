import openpyxl
import numpy

wb = openpyxl.load_workbook('DDY.xlsx')
data_sheet = wb['данные за 2 кв.']

classes = {}
room_numbers = {}
for i in range(2, data_sheet.max_row + 1):
    classes[data_sheet.cell(row=i, column=53).value] = [0, [], []]
    room_numbers[data_sheet.cell(row=i, column=11).value] = [0, [], []]
for i in range(2, data_sheet.max_row + 1):
    if data_sheet.cell(row=i, column=53):
        classes[data_sheet.cell(row=i, column=53).value][0] += 1
        classes[data_sheet.cell(row=i, column=53).value][1].append(data_sheet.cell(row=i, column=10).value)
        if i not in [5, 6, 7, 8]:
            classes[data_sheet.cell(row=i, column=53).value][2].append(data_sheet.cell(row=i, column=27).value)
    if data_sheet.cell(row=i, column=11) and str(data_sheet.cell(row=i, column=27).value) != 'None':
        room_numbers[data_sheet.cell(row=i, column=11).value][0] += 1
        room_numbers[data_sheet.cell(row=i, column=11).value][1].append(data_sheet.cell(row=i, column=10).value)
        if i not in [5, 6, 7, 8]:
            room_numbers[data_sheet.cell(row=i, column=11).value][2].append(data_sheet.cell(row=i, column=27).value)
room_s = {}
for i in range(2, data_sheet.max_row + 1):
    room_s[str(data_sheet.cell(row=i, column=53).value) + ' ' + str(data_sheet.cell(row=i, column=11).value)] = [0, [], []]
for i in range(2, data_sheet.max_row + 1):
    if data_sheet.cell(row=i, column=53) and data_sheet.cell(row=i, column=11):
        if str(data_sheet.cell(row=i, column=11).value) != 'None' and str(data_sheet.cell(row=i, column=53).value) != 'Нет такого корпуса':
            room_s[str(data_sheet.cell(row=i, column=53).value) + ' ' + str(data_sheet.cell(row=i, column=11).value)][0] += 1
            room_s[str(data_sheet.cell(row=i, column=53).value) + ' ' + str(data_sheet.cell(row=i, column=11).value)][1].append(data_sheet.cell(row=i, column=10).value)
            if i not in [5, 6, 7, 8]:
                room_s[str(data_sheet.cell(row=i, column=53).value) + ' ' + str(data_sheet.cell(row=i, column=11).value)][2].append(data_sheet.cell(row=i, column=27).value)

print(f'Классы: {classes}')
print(f'Комнатность: {room_numbers}')
print(f'Комнатность по классам: {room_s}')

new_wb = openpyxl.Workbook()
n_sheet = new_wb.active
n_sheet.title = 'Лист1'

i = 2
n_sheet['A1'] = 'Комнатность'
n_sheet['B1'] = 'Куплено штук'
n_sheet['C1'] = 'Средняя площадь'
n_sheet['D1'] = 'Средняя цена'
for key, value in room_numbers.items():
    if key is None:
        n_sheet.cell(row=i, column=1).value = 'Нет данных'
        n_sheet.cell(row=i, column=2).value = value[0]
        n_sheet.cell(row=i, column=3).value = str(round(numpy.mean(value[1]), 1))
        n_sheet.cell(row=i, column=4).value = str(round(numpy.mean((value[2])), 1))
    else:
        n_sheet.cell(row=i, column=1).value = key
        n_sheet.cell(row=i, column=2).value = value[0]
        n_sheet.cell(row=i, column=3).value = str(round(numpy.mean(value[1]), 1))
        n_sheet.cell(row=i, column=4).value = str(round(numpy.mean(value[2]), 1))
    i += 1

j = 2
n_sheet['F1'] = 'Класс'
n_sheet['G1'] = 'Куплено штук'
n_sheet['H1'] = 'Средняя площадь'
n_sheet['I1'] = 'Средняя цена'
for key, value in classes.items():
    if key == 'Нет такого корпуса':
        n_sheet.cell(row=j, column=6).value = 'Нет данных'
        n_sheet.cell(row=j, column=7).value = value[0]
        n_sheet.cell(row=j, column=8).value = str(round(numpy.mean(value[1]), 1))
        n_sheet.cell(row=j, column=9).value = str(round(numpy.mean(value[2]), 1))
    else:
        n_sheet.cell(row=j, column=6).value = key
        n_sheet.cell(row=j, column=7).value = value[0]
        n_sheet.cell(row=j, column=8).value = str(round(numpy.mean(value[1]), 1))
        n_sheet.cell(row=j, column=9).value = str(round(numpy.mean(value[2]), 1))
    j += 1

k = 2
n_sheet['K1'] = 'Класс + комнатность'
n_sheet['L1'] = 'Куплено штук'
n_sheet['M1'] = 'Средняя площадь'
n_sheet['N1'] = 'Средняя цена'
for key, value in room_s.items():
    if value[0] != 0 and key:
        n_sheet.cell(row=k, column=11).value = key
        n_sheet.cell(row=k, column=12).value = str(value[0])
        n_sheet.cell(row=k, column=13).value = str(round(numpy.mean(value[1]), 1))
        n_sheet.cell(row=k, column=14).value = str(round(numpy.mean(value[2]), 1))
        k += 1

new_wb.save('analytics.xlsx')
