import openpyxl
import numpy

wb = openpyxl.load_workbook('DDY.xlsx')
data_sheet = wb['данные за 2 кв.']

names = {}

for i in range(2, data_sheet.max_row):
    names[data_sheet.cell(row=i, column=2).value] = [0, [], []]

for i in range(2, data_sheet.max_row):
    if data_sheet.cell(row=i, column=2).value in names:
        names[data_sheet.cell(row=i, column=2).value][0] += 1
        if data_sheet.cell(row=i, column=27).value is not None:
            names[data_sheet.cell(row=i, column=2).value][1].append(data_sheet.cell(row=i, column=27).value)
        if str(data_sheet.cell(row=i, column=11).value).isdigit():
            names[data_sheet.cell(row=i, column=2).value][2].append(int(data_sheet.cell(row=i, column=11).value))

sells = set()
for key, value in names.items():
    names[key].append(len(value[1]))
    sells.add(len(value[1]))

sells = list(sells)

top1 = {}
def top(arr):
    for key, value in names.items():
        if value[3] == max(arr):
            top1[key] = value[3]
            top(arr[:-2])
        if len(arr) == 1:
            break

top(sells)
top_sort = []
for key, value in top1.items():
    top_sort = sorted(top1.values())
top_sort = top_sort[::-1]
top_sort = top_sort[:5]
for key, value in top1.items():
    if value in top_sort:
        print(f'{key}: {value}')

print(names)
print(top_sort)

for key, value in names.items():
    if value[3] in top_sort:
        print(f'{key}: {value[3]}: {round(numpy.mean(value[1]), 1)}')
for key, value in names.items():
    if value[2] and value[3] in top_sort:
        print(f'{key}: {value[3]}: {round(numpy.mean(value[1]), 1)}: {value[2].count(9)}')


