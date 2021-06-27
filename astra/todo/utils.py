import csv

from openpyxl import Workbook, load_workbook


def export_to_csv(task, fields, file_name):
    # with open(f'{file_name}.csv', "w", newline='', encoding='utf-8') as csv_file:
    #     writer = csv.writer(csv_file, delimiter=';')
    #
    #     if fields:
    #         headers = fields
    #         if titles:
    #             titles = titles
    #         else:
    #             titles = headers
    #     else:
    #         headers = []
    #         titles = headers
    #
    #     writer.writerow(titles)
    #     data_to_export = [task.name, task.deadline, task.description, task.done, task.created_at, task.updated_at, task.is_active]
    #     writer.writerow(data_to_export)
    wb = Workbook()
    ws = wb.active
    ws.title = "Task"
    wb.save(filename=f'{file_name}.xlsx')

    filename = f'{file_name}.xlsx'
    wb = load_workbook(filename)
    file = wb.active

    file.append(fields)
    file.append([task.name, task.deadline, task.description, task.done, task.created_at, task.updated_at, task.is_active])

    wb.save(filename=filename)
