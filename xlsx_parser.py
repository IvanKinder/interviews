import openpyxl

wb = openpyxl.load_workbook('DDY.xlsx')

data_sheet = wb['данные за 2 кв.']
classes_sheet = wb['lib класс']
rooms_sheet = wb['lib комнатность']

id_in_data = {}
for i in range(2, data_sheet.max_row):
    if data_sheet.cell(row=i, column=32).value == '':
        continue
    id_in_data[data_sheet.cell(row=i, column=32).value] = (
    data_sheet.cell(row=i, column=1).value, data_sheet.cell(row=i, column=34).value,
    data_sheet.cell(row=i, column=11).value, data_sheet.cell(row=i, column=10).value, [])

id_in_classes = {}
for i in range(2, classes_sheet.max_row):
    if classes_sheet.cell(row=i, column=3).value == '':
        continue
    id_in_classes[classes_sheet.cell(row=i, column=3).value] = (
    classes_sheet.cell(row=i, column=1).value, classes_sheet.cell(row=i, column=4).value)

id_in_rooms = []
for i in range(2, rooms_sheet.max_row):
    id_in_rooms.append([rooms_sheet.cell(row=i, column=1).value, rooms_sheet.cell(row=i, column=2).value,
                        rooms_sheet.cell(row=i, column=3).value, rooms_sheet.cell(row=i, column=4).value])

wb.active = wb['данные за 2 кв.']

data_sheet['BA1'].value = 'Верный класс'
for i in range(2, data_sheet.max_row + 1):
    if data_sheet.cell(row=i, column=32).value not in id_in_classes.keys():
        data_sheet.cell(row=i, column=53).value = 'Нет такого корпуса'
    elif id_in_data[data_sheet.cell(row=i, column=32).value][:1] == id_in_classes[
        data_sheet.cell(row=i, column=32).value]:
        data_sheet.cell(row=i, column=53).value = id_in_classes[data_sheet.cell(row=i, column=32).value][1]
    else:
        data_sheet.cell(row=i, column=53).value = id_in_classes[data_sheet.cell(row=i, column=32).value][1]

for value in id_in_data.values():
    if value[2]:
        for room in id_in_rooms:
            if (str(room[1]) == value[2] and room[2] <= value[3] <= room[3]) or (
                    str(room[1]) == '4+' and value[2] == '5'):
                value[4].append(room[0])

data_sheet['BB1'].value = 'Верная комнатность'
for i in range(2, data_sheet.max_row + 1):
    if id_in_data[data_sheet.cell(row=i, column=32).value][4]:
        if len(id_in_data[data_sheet.cell(row=i, column=32).value][4]) == 1:
            data_sheet.cell(row=i, column=54).value = ', '.join(id_in_data[data_sheet.cell(row=i, column=32).value][4])
        else:
            data_sheet.cell(row=i, column=54).value = 'Пересечение данных: ' + ', '.join(
                id_in_data[data_sheet.cell(row=i, column=32).value][4])
    else:
        if data_sheet.cell(row=i, column=11).value:
            data_sheet.cell(row=i, column=54).value = 'Нет ID Корпуса или данных по такой площади'
        else:
            data_sheet.cell(row=i, column=54).value = 'Не хватает данных (комнатности)'

wb.save('DDY.xlsx')
